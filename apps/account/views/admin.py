#!/usr/bin/env python
# coding: utf-8

import logging
import os
import shutil

import magic
import openpyxl
from django.conf import settings
from django.db import IntegrityError
from django.db.models import Q
from django.http import HttpResponse

from apps.account.decorators import super_admin_required, manager_admin_required
from apps.account.models import AdminType, CasePermission, User, UserProfile
from apps.account.serializers import (
    ImageUploadForm,
    ImportUserSeralizer,
    UserAdminSerializer,
)
from apps.utils.api import APIView, validate_serializer
from apps.utils.api.api import CSRFExemptAPIView
from apps.utils.client import user_manager_client
from apps.utils.handle_passwd import encode_passwd
from apps.utils.shortcuts import rand_str

logger = logging.getLogger(__name__)


class UserAdminAPI(APIView):
    @manager_admin_required
    def get(self, request):
        """
        获取用户列表或者通过用户 ID 来获取用户
        """
        user_id = request.GET.get("id")
        if user_id:
            try:
                user = User.objects.get(id=user_id)
            except User.DoesNotExist:
                return self.error("User does not exist")
            return self.success(UserAdminSerializer(user).data)

        user = User.objects.exclude(id=1).order_by("-create_time")

        # 按照关键字查询
        keyword = request.GET.get("keyword", None)
        if keyword:
            user = user.filter(
                Q(username__icontains=keyword)
                | Q(userprofile__real_name__icontains=keyword)
                | Q(email__icontains=keyword)
            ).distinct()
        return self.success(self.paginate_data(request, user, UserAdminSerializer))

    @validate_serializer(ImportUserSeralizer)
    @super_admin_required
    def post(self, request):
        pass

    @super_admin_required
    def put(self, request):
        """
        修改用户信息的 API。
        """
        data = request.data

        try:
            user = User.objects.get(id=data["id"])
        except User.DoesNotExist:
            return self.error("User does not exist")
        if User.objects.filter(username=data["username"]).exclude(id=user.id).exists():
            return self.error("Username already exists")

        # 获取数据后直接更改
        pre_username = user.username
        print(pre_username)  # dummy info
        user.username = data["username"]
        user.email = data["email"]
        user.admin_type = data["admin_type"]
        user.is_disabled = data["is_disabled"]
        new_password = data["password"]

        if data["admin_type"] == AdminType.ADMIN:
            user.case_permission = CasePermission.OWN
        elif (
            data["admin_type"] == AdminType.SUPER_ADMIN
            or data["admin_type"] == AdminType.INTERNAL_ADMIN
        ):
            user.case_permission = CasePermission.ALL
        else:
            user.case_permission = CasePermission.NONE

        form_data = {
            "pre_username": pre_username,
            "username": data["username"],
            "new_password": new_password,
        }
        msg = user_manager_client.modify(form_data)
        if msg:
            return self.error(msg)
        if new_password:
            user.set_password(new_password)
            encrypt_passwd = encode_passwd(new_password)
            user.encrypt_passwd = encrypt_passwd
        user.save()

        # 更新对应的用户详情
        UserProfile.objects.filter(user=user).update(real_name=data["real_name"])
        return self.success()

    @super_admin_required
    def delete(self, request):
        pass


class DleteServerUsers(APIView):
    @super_admin_required
    def delete(self, request):
        """
        删除用户的同时删除其他系统用户
        """
        names = request.GET.get("user")
        if not names:
            return self.error("Invalid Parameter, name is required")
        user_names = names.split(",")
        data = user_manager_client.delete(names)
        if msg := data.get("msg"):
            return self.error(msg)
        for user_name in user_names:
            User.objects.filter(username=user_name).delete()
        return self.success(data)


class UploadRegisterData(CSRFExemptAPIView):
    """
    上传批量注册用户的xlsx文件
    """

    request_parsers = ()

    @super_admin_required
    def post(self, request):
        files = request.FILES.get("file")
        fileName = files.name

        if not fileName.endswith("xlsx"):
            error_msg = "Error, file format error, expected xlsx file"
            logger.error(error_msg)
            return self.error(error_msg)

        if len(files) == 0:
            error_msg = "Error, no file found !"
            logger.error(error_msg)
            return self.error(error_msg)

        tempDir = os.path.join(settings.TMP_DIR, rand_str())
        if os.path.exists(tempDir):
            shutil.rmtree(tempDir)
        os.makedirs(tempDir, exist_ok=True)

        uploadFile = os.path.join(tempDir, "data.xlsx")
        # 创建子目录
        os.makedirs(os.path.dirname(uploadFile), exist_ok=True)
        # 写入数据
        with open(uploadFile, "wb") as f:
            for chunk in files:
                f.write(chunk)

        # manual_file_id, info = process_uploadFiles(settings.REGISTER_DIR, uploadFile)  # noqa
        os.makedirs(settings.REGISTER_DIR, exist_ok=True)
        shutil.copy(uploadFile, os.path.join(settings.REGISTER_DIR, "data.xlsx"))
        shutil.rmtree(tempDir)

        return self.success({"info": "success"})


class GenerateUserAPI(CSRFExemptAPIView):
    @super_admin_required
    def get(self, request):
        keyword = request.GET.get("key")
        if not keyword:
            file_path = os.path.join(settings.REGISTER_DIR, "data.xlsx")
        else:
            current_path = os.path.dirname(os.path.dirname(__file__))
            templates_path = os.path.join(current_path, "templates")
            file_path = os.path.join(templates_path, "template.xlsx")
        try:
            workbook = openpyxl.load_workbook(file_path)
            response = HttpResponse(content_type="application/octet-stream")
            response["content-Disposition"] = "attachment;filename=data.xlsx"
            workbook.save(response)
            return response
        except:
            return self.error(msg="download error")

    @super_admin_required
    def post(self, request):
        file_path = os.path.join(settings.REGISTER_DIR, "data.xlsx")
        # try to load xlsx
        try:
            workbook = openpyxl.load_workbook(file_path)
            worksheet = workbook.worksheets[0]
            assert worksheet.cell(1, 1).value.strip() == "用户名"
            assert worksheet.cell(1, 2).value.strip() == "密码"
        except:
            os.remove(file_path)
            return self.error(msg="error", err="请检查是否存在sheet、以及从左上开始是否为（A1用户名、B1密码）")

        try:
            workbook = openpyxl.load_workbook(file_path)
            worksheet = workbook.worksheets[0]
            user_list = []
            row_num = 2
            while True:
                username = worksheet.cell(row_num, 1).value
                raw_password = worksheet.cell(row_num, 2).value
                if not username:
                    break
                if not raw_password:
                    raw_password = rand_str(length=8)
                    worksheet.cell(row_num, 3, value=raw_password)
                user = {
                    "username": str(username),
                    "password": str(raw_password),
                }
                user_list.append(user)
                row_num += 1
            workbook.save(filename=file_path)
        except Exception as e:
            os.remove(file_path)
            return self.error(msg="error", err=str(e))

        try:
            for user_info in user_list:
                # 创建远程试用系统用户
                user = User.objects.create(username=user_info.get("username"))
                user.set_password(user_info.get("password"))
                user.save()
                # 创建用户详情
                UserProfile.objects.create(user=user)
                # 创建关联服务用户
                form_data = {
                    "user_name": user_info.get("username"),
                    "user_passwd": user_info.get("password"),
                }
                msg = user_manager_client.register(form_data)
                if msg:
                    # 删除远程试用系统用户
                    User.objects.filter(username=user_info.get("username")).delete()
                    return self.error(msg)
            return self.success({"info": "success"})

        except IntegrityError as e:
            os.remove(file_path)
            return self.error(str(e).split("\n")[1])


class UploadLogoAPI(APIView):
    request_parsers = ()

    def post(self, request):
        """
        上传系统Logo图片
        """
        form = ImageUploadForm(request.POST, request.FILES)
        if form.is_valid():
            logo = form.cleaned_data["image"]
        else:
            return self.error("Invalid file content")
        # 图像大小限制
        if logo.size > 2 * 1024 * 1024:
            return self.error("Picture is too large")
        # 图像格式的限制
        tempImage = os.path.join(settings.TMP_DIR, "tmpImg")
        with open(tempImage, "wb") as tempImg:
            for chunk in logo:
                tempImg.write(chunk)
        upFileType = magic.from_file(tempImage, mime=True)
        if upFileType not in [
            "image/gif",
            "image/jpeg",
            "image/png",
            "image/bmp",
            "image/svg+xml",
        ]:
            return self.error("Unsupported file format")
        os.remove(tempImage)

        if not os.path.exists(settings.LOGO_UPLOAD_DIR):
            os.mkdir(settings.LOGO_UPLOAD_DIR)
        logoFileName = os.path.join(settings.LOGO_UPLOAD_DIR, "logo.png")
        with open(logoFileName, "wb") as img:
            for chunk in logo:
                img.write(chunk)
        return self.success("Succeeded")


class SynchronizeAPI(APIView):
    """
    同步用户
    """

    @super_admin_required
    def put(self, request):
        # 获取所有用户
        user = User.objects.exclude(id=1)
        form_data = {}
        for single_user in user:
            form_data.update({single_user.username: single_user.encrypt_passwd})
        data = user_manager_client.update(form_data)
        if msg := data.get("msg"):
            return self.error(msg)
        return self.success(data)
